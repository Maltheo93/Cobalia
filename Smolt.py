import csv
import glob
import json
import locale
import os
import re
import shutil
import time
from datetime import datetime, timedelta, date

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait

# Set up the driver with notifications disabled, downloads allowed and then navigate to Cobália
options = Options()
options.add_argument("--disable-notifications")
options.add_argument("--safebrowsing-disable-download-protection")
prefs = {
    "download.default_directory": os.path.join(os.path.expanduser("~"), "Downloads"),
    "download.prompt_for_download": False,
    "safebrowsing.enabled": False,
    "safebrowsing.disable_download_protection": True,
    "profile.default_content_settings.popups": 0,
    "profile.default_content_setting_values.automatic_downloads": 1,
}
options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.get("https://www.cobalia.com")

# Load JSON data with danish encoding
with open(
        'C:/Users/mamo/Desktop/Dambrug Monthly Reports/Facility Info.JSON', 'r',
        encoding='utf-8') as f:
    json_data = json.load(f)

# Extract selectors for all elements as well as facility and directory information
element_selectors = {}
facilities = []
directories = {}

for element_name, element_data in json_data['elements'].items():
    element_selectors[element_name] = element_data['selector']

for facility in json_data['facilities']:
    facility_id = facility['id']
    facility_name = facility['facility_name']
    cobalia_id = facility['cobalia_id']
    facility_directory = facility['facility_directory']
    facilities.append(
        {
            'id': facility_id,
            'facility_name': facility_name,
            'cobalia_id': cobalia_id,
            'facility_directory': facility_directory
        })

directories['download_path'] = json_data['directories']['download_path']
directories['smolt_report_path'] = json_data['directories']['smolt_report_path']


# Set up functions

# Click function to click elements
# Use Example: click_elements(driver, {"username_input": element_selectors["username_input"]})
def click_elements(driver, selectors):
    for element_name, selector in selectors.items():
        try:
            element = WebDriverWait(driver, 5).until(
                ec.element_to_be_clickable((By.XPATH, selector)))
            element.click()
        except TimeoutException:
            print(f"Error: Timed out waiting for selector: {element_name}")
        except Exception as e:
            print(f"Error: {e} occurred while clicking selector: {element_name}")

# Reset calendar function
# Use Example: reset_calendar(driver, element_selectors)
def reset_calendar(driver, element_selectors):
    try:
        time.sleep(1)
        click_elements(
            driver, {
                "calendar_one_1": element_selectors["calendar_one"],
                "calendar_second": element_selectors["calendar_second"],
                "calendar_one_2": element_selectors["calendar_one"],
                "calendar_first": element_selectors["calendar_first"]
            })
        time.sleep(1)
    except TimeoutException:
        print("Error: TimeoutException occurred while resetting calendar")


# Download PDF function
# Use Example: download_pdf(driver, element_selectors)
def download_pdf(driver, element_selectors):
    click_elements(
        driver, {
            "download_dropdown": element_selectors["download_dropdown"],
            "pdf_select": element_selectors["pdf_select"],
            "pdf_print": element_selectors["pdf_print"],
            "pdf_download": element_selectors["pdf_download"]
        })


# Download CSV function
# Use Example: download_csv(driver, element_selectors)
def download_csv(driver, element_selectors):
    click_elements(
        driver, {
            "download_dropdown": element_selectors["download_dropdown"],
            "csv_select": element_selectors["csv_select"],
            "csv_download": element_selectors["csv_download"]
        })


# Delete the newest file in the directory that contains "Facilitetsrapport" in the filename and has a .csv extension
# Use Example: delete_csv(directories['download_path'])
def delete_csv(download_path):
    time.sleep(3)
    file_pattern = os.path.join(download_path, "*Facilitetsrapport*.csv")
    files = glob.glob(file_pattern)
    if len(files) > 0:
        newest_file = max(files, key=os.path.getctime)
        os.remove(newest_file)


# Function to move and rename CSV and PDF files to correct facility-corresponding paths
# Use example: move_monthly_reports(facility_id, facilities)
def move_monthly_reports(facility_id, facilities, directories):
    # Calculate previous month and specify source and destination folders
    now = datetime.now()
    previous_month = now.replace(day=1) - timedelta(days=1)
    previous_month_str = previous_month.strftime("%m-%Y")

    facility = None
    for f in facilities:
        if f['id'] == facility_id:
            facility = f
            break

    if facility is None:
        print(f"Error: facility with id {facility_id} not found")
        exit()
    facility_name = facility["facility_name"]
    facility_directory = facility["facility_directory"]

    # Get the list of csv and pdf files in the source folder, sorted by modification time
    download_path = directories['download_path']
    csv_files = glob.glob(os.path.join(download_path, "*.csv"))
    pdf_files = glob.glob(os.path.join(download_path, "*.pdf"))
    csv_files.sort(key=os.path.getmtime)
    pdf_files.sort(key=os.path.getmtime)

    # Get the name of the newest csv and pdf files, and construct new names before moving to destination
    newest_csv_file = csv_files[-1]
    newest_pdf_file = pdf_files[-1]
    new_csv_file_name = f"{previous_month_str} {facility_name}.csv"
    new_pdf_file_name = f"{previous_month_str} {facility_name}.pdf"
    shutil.move(newest_csv_file, os.path.join(facility_directory, new_csv_file_name))
    shutil.move(newest_pdf_file, os.path.join(facility_directory, new_pdf_file_name))


# Function to update smolt report excel file
# Use Example: update_smolt_report(facility_id, facilities, directories)
def update_smolt_report(facility_id, facilities, directories):
    # Get the facility name from the facilities list using the facility_id
    facility = None
    for f in facilities:
        if f['id'] == facility_id:
            facility = f
            break

    if facility is None:
        print(f"Error: facility with id {facility_id} not found")
        exit()
    facility_name = facility["facility_name"]

    # Set up paths and names, then copy and rename file
    source_file_name = "Skabelon.xlsx"
    today = date.today()
    prev_month = today.replace(day=1) - timedelta(days=1)
    facility = None
    for f in facilities:
        if f['id'] == facility_id:
            facility = f
            break

    if facility is None:
        print(f"Error: facility with id {facility_id} not found")
        exit()

    facility_path = facility["facility_directory"]
    dest_file_name = "{:02d}-{} {}.xlsx".format(
        prev_month.month, prev_month.year, facility_name)
    smolt_report_path = directories["smolt_report_path"]
    shutil.copy(os.path.join(smolt_report_path, source_file_name), facility_path)
    os.rename(
        os.path.join(facility_path, source_file_name),
        os.path.join(facility_path, dest_file_name))

    # Set up file paths and names
    folder_name = os.path.basename(facility_path)
    previous_month = (date.today().replace(day=1) - timedelta(days=1)).replace(day=1)
    prev_month_name_da = \
        "Januar Februar Marts April Maj Juni Juli August September Oktober November December".split()[
            previous_month.month - 1]
    prev_month_year = previous_month.year

    # Load the XLSX file and get a reference to the existing sheet
    destination_file = os.path.join(
        facility_path, f'{previous_month.strftime("%m-%Y")} {folder_name}.xlsx')
    book = openpyxl.load_workbook(destination_file)
    sheet = book['Data']

    # Read the data from the CSV file into a list of lists
    with open(
            os.path.join(
                facility_path, f'{previous_month.strftime("%m-%Y")} {folder_name}.csv'),
            newline='', encoding='utf-8-sig') as csvfile:
        reader = csv.reader(csvfile, delimiter='\t')
        data = list(reader)
    locale.setlocale(locale.LC_ALL, 'da_DK.UTF-8')

    # Loop through the list of lists and write the data to the sheet
    for row_index, row_data in enumerate(data):
        for col_index, col_data in enumerate(row_data):
            # Check if the data is a number
            if re.match(r'^[0-9,.]+$', col_data):
                # Convert the string to a float using the Danish number system
                col_data = locale.atof(col_data)
                # Set the cell format to be numeric
                sheet.cell(
                    row=row_index + 4, column=col_index + 1).number_format = '#,##0.00'
            sheet.cell(row=row_index + 4, column=col_index + 1).value = col_data

    # Update Summation sheet
    summation_sheet = book['Summation']
    summation_sheet['C4'] = folder_name
    summation_sheet['C6'] = prev_month_name_da
    summation_sheet['C8'] = prev_month_year

    book.save(destination_file)
    book.close()


# Log into Cobália
login_data = json_data['login']
username = login_data['username']
password = login_data['password']

username_input_selector = element_selectors["username_input"]
username_input = WebDriverWait(driver, 5).until(
    ec.element_to_be_clickable((By.XPATH, username_input_selector)))
username_input.send_keys(username)

password_input_selector = element_selectors["password_input"]
password_input = WebDriverWait(driver, 5).until(
    ec.element_to_be_clickable((By.XPATH, password_input_selector)))
password_input.send_keys(password)

click_elements(driver, {"login_button": element_selectors["login_button"]})

# Navigate to the facility report page and set Calendar 1 to the first day of last month
driver.get("https://www.cobalia.com/operation/facility")
time.sleep(0.5)
click_elements(
    driver, {
        "date_range": element_selectors["date_range"],
        "user_defined": element_selectors["user_defined"],
        "calendar_one": element_selectors["calendar_one"],
        "calendar_oneprev": element_selectors["calendar_oneprev"],
        "calendar_first": element_selectors["calendar_first"]
    })

# Set Calendar 2 to the last day of last month
time.sleep(0.25)
click_elements(
    driver, {
        "calendar_two": element_selectors["calendar_two"],
        "calendar_last": element_selectors["calendar_last"]
    })
for i in range(31, 27, -1):
    day_xpath = f'//span[@class="flatpickr-day" and @aria-label and not(contains(@class, "prevMonthDay") or contains(@class, "nextMonthDay"))][{i}]'
    try:
        day_element = WebDriverWait(driver, 0.5).until(
            ec.element_to_be_clickable((By.XPATH, day_xpath)))
        day_element.click()
        break
    except TimeoutException:
        if i == 28:
            break
        continue

# Loop through facility IDs 1 to 10 and perform all actions for each one
for facility_id in range(1, 11):
    # Get the facility data from the JSON data
    facility_data = facilities[facility_id - 1]
    facility_name = facility_data["facility_name"]
    facility_name_selector = element_selectors[facility_name]
    time.sleep(1)
    click_elements(driver, {"user_dropdown": element_selectors["user_dropdown"]})
    if facility_id == 1:
        click_elements(driver, {"facility_dropdown": element_selectors["facility_dropdown"]})
    time.sleep(1)
    click_elements(driver, {"facility_name": facility_name_selector})
    time.sleep(1)
    reset_calendar(driver, element_selectors)
    time.sleep(1)
    download_pdf(driver, element_selectors)
    time.sleep(1)
    download_csv(driver, element_selectors)
    time.sleep(1)
    delete_csv(directories['download_path'])
    time.sleep(1)
    directories = json_data['directories']
    move_monthly_reports(facility_id, facilities, directories)
    time.sleep(1)
    update_smolt_report(facility_id, facilities, directories)

time.sleep(10)
